import json
import re
from datetime import date, timedelta
from pathlib import Path

import openpyxl


ROOT = Path(r"D:\Jerry's Test Folder\sportsmedals-main")
DATA = ROOT / "data"


def parse_time(text):
  s = str(text or "").strip()
  if not s:
    return None, None, None
  day = None
  slot = None
  venue = ""
  day_match = re.search(r"D\s*(\d+)", s, re.I)
  if day_match:
    day = int(day_match.group(1))
  slot_match = re.search(r"G\s*(\d+)", s, re.I)
  if slot_match:
    index = int(slot_match.group(1))
    slot = "" if index <= 0 else f"G{index}"
  venue_match = re.search(r"\b([A-C]\d+)\b", s)
  if venue_match:
    venue = venue_match.group(1)
  return day, slot, venue


def clean_name(raw):
  return str(raw or "").replace("比赛项目：", "").strip().replace(" ", "")


def simplify_team_name(raw):
  text = str(raw or "").strip()
  if not text:
    return ""
  text = text.replace("班", "")
  text = re.sub(r"\s+", " ", text).strip()

  spaced = re.match(r"^(中预|初[一二三]|高[一二三])\s*(\d{1,2})\s+(\d{1,2})$", text)
  if spaced:
    return f"{spaced.group(1)}{spaced.group(2)} 组{spaced.group(3)}"

  compact = re.sub(r"\s+", "", text)
  compact = re.sub(r"^(中预|初[一二三]|高[一二三])(\d{1,2})组(\d{1,2})$", r"\1\2 组\3", compact)
  compact = re.sub(r"^(中预|初[一二三]|高[一二三])(\d{1,2})(男|女)(\d{1,2})$", r"\1\2 \3\4", compact)
  return compact


def clean_grade(raw):
  return str(raw or "").replace("比赛年级：", "").strip()


def stage_name(raw):
  text = str(raw or "").replace("比赛阶段：", "").strip()
  if not text:
    return "一"
  return "一" if text == "0" else text


def build_day_map(max_day):
  base_date = date(2026, 3, 23)
  day_map = {}
  current = base_date
  day = 1
  while day <= max_day:
    if current.weekday() < 5:
      day_map[str(day)] = f"{current.month}.{current.day}"
      day += 1
    current += timedelta(days=1)
  for day in range(1, max_day + 1):
    current = day_map.get(str(day))
    if current:
      continue
    current_date = base_date + timedelta(days=day - 1)
    day_map[str(day)] = f"{current_date.month}.{current_date.day}"
  return day_map


def build_default_time_slots(max_g=7):
  slots = {}
  if max_g <= 0:
    return slots

  def fmt(minutes):
    h = int(minutes // 60)
    m = int(round(minutes % 60))
    return f"{h:02d}:{m:02d}"

  for index in range(1, max_g + 1):
    late_s = 17 * 60 + (index - 1) * 30
    late_e = late_s + 30
    noon_s = 13 * 60 + 30 + (index - 1) * 30
    noon_e = noon_s + 30
    slots[f"G{index}"] = f"{fmt(late_s)}-{fmt(late_e)}"
    slots[f"D5_G{index}"] = f"{fmt(noon_s)}-{fmt(noon_e)}"
    slots[f"D10_G{index}"] = f"{fmt(noon_s)}-{fmt(noon_e)}"

  return slots


def section_key(sheet, project, grade):
  suffix = ""
  if "男" in grade:
    suffix = "-male"
  elif "女" in grade:
    suffix = "-female"

  if sheet == "足球":
    return "football", "足球", True
  if sheet == "排球":
    return "volleyball", project or "排球", False
  if sheet == "乒乓球":
    return "table-tennis", project or "乒乓球", False
  if sheet == "混双":
    return "badminton-mixed", "羽毛球混双", False
  if sheet == "网球":
    if suffix == "-male":
      return "tennis-male", "网球男单", False
    if suffix == "-female":
      return "tennis-female", "网球女单", False
    return "tennis", "网球单打", False
  if sheet == "篮球":
    if "3v3" in project:
      if suffix == "-male":
        return "basketball-3v3-male", "篮球 3v3 男", True
      if suffix == "-female":
        return "basketball-3v3-female", "篮球 3v3 女", True
      return "basketball-3v3", "篮球 3v3", True
    if "技巧" in project:
      return "basketball-skill", "篮球技巧赛", False
    if "投篮" in project:
      return "basketball-shooting", "投篮挑战", False
    return "basketball", "篮球", True

  base = re.sub(r"[^a-z0-9-]+", "-", sheet.lower())
  return base, project or sheet, False


def stage_gender_suffix(grade):
  if "男" in grade:
    return "（男）"
  if "女" in grade:
    return "（女）"
  return ""


def normalize_sport_id(base, grade_tag, project):
  if base == "basketball-3v3-male":
    return f"basketball-male-{grade_tag}"
  if base == "basketball-3v3-female":
    return f"basketball-female-{grade_tag}"
  if base == "basketball-3v3":
    return f"basketball-{grade_tag}"
  if base == "basketball-shooting":
    return f"basketball-skill-{grade_tag}"
  return f"{base}-{grade_tag}"


def parse_workbook(path, grade_tag):
  workbook = openpyxl.load_workbook(path, data_only=True)
  sports_map = {}
  match_seq = 1

  for worksheet in workbook.worksheets:
    row_index = 1
    while row_index <= worksheet.max_row:
      start_value = worksheet.cell(row_index, 1).value
      if not (isinstance(start_value, str) and "比赛项目" in start_value):
        row_index += 1
        continue

      project = clean_name(worksheet.cell(row_index, 1).value)
      grade = clean_grade(worksheet.cell(row_index, 3).value)
      stage = stage_name(worksheet.cell(row_index, 5).value)

      sport_id_base, sport_name, team_event = section_key(worksheet.title, project, grade)
      sport_id = normalize_sport_id(sport_id_base, grade_tag, project)

      if sport_id not in sports_map:
        sports_map[sport_id] = {
          "id": sport_id,
          "name": f"{grade_tag.upper()} · {sport_name}",
          "teamEvent": team_event,
          "teamMap": {},
          "stages": []
        }
      sport = sports_map[sport_id]

      sport["name"] = sport_name

      stage_label = stage
      suffix = stage_gender_suffix(grade)
      if suffix:
        stage_label = f"{stage}{suffix}"

      section_end = row_index + 1
      while section_end <= worksheet.max_row:
        current = worksheet.cell(section_end, 1).value
        if isinstance(current, str) and "比赛项目" in current:
          break
        section_end += 1

      for read_row in range(row_index, section_end):
        for read_col in range(8, 14):
          mapping_value = worksheet.cell(read_row, read_col).value
          if not mapping_value:
            continue
          mapping_text = str(mapping_value).strip()
          code_match = re.match(r"^([A-Z])\s*(.+)$", mapping_text)
          if code_match:
            sport["teamMap"][code_match.group(1)] = simplify_team_name(code_match.group(2))

      matches = []
      for read_row in range(row_index + 1, section_end):
        for left_col in (1, 3, 5):
          code_value = worksheet.cell(read_row, left_col).value
          time_value = worksheet.cell(read_row, left_col + 1).value
          if not code_value and not time_value:
            continue

          code_text = str(code_value or "").strip()
          time_text = str(time_value or "").strip()

          if not time_text and re.search(r"D\s*\d+", code_text, re.I):
            time_text = code_text
            code_text = ""

          if not code_text and time_text and not team_event:
            code_text = project
          if not code_text:
            continue

          day, slot, venue = parse_time(time_text)
          day_value = day if day is not None else 1

          match_id = f"{sport_id}-{match_seq}"
          match_seq += 1

          if "/" in code_text:
            match_no = code_text.split("/", 1)[0].strip()
          else:
            match_no = str(len(matches) + 1)

          matches.append({
            "id": match_id,
            "matchNo": match_no,
            "code": code_text,
            "day": day_value,
            "slot": slot or "",
            "venue": venue,
            "teams": [
              {"code": "", "name": ""},
              {"code": "", "name": ""}
            ]
          })

          participants = []
          if "/" in code_text:
            payload = code_text.split("/", 1)[1].strip()
            star_parts = [part.strip() for part in payload.split("*") if part.strip()]
            if len(star_parts) >= 2:
              participants = [star_parts[0], star_parts[1]]
            elif re.fullmatch(r"[A-Z]{2}", payload):
              participants = [payload[0], payload[1]]
            else:
              token_parts = re.findall(r"\d+U\d?|\d+[WL]|[A-Z]", payload)
              if len(token_parts) >= 2:
                participants = [token_parts[0], token_parts[1]]

          if participants:
            matches[-1]["teams"][0]["code"] = participants[0]
            matches[-1]["teams"][1]["code"] = participants[1]
            if participants[0] in sport["teamMap"]:
              matches[-1]["teams"][0]["name"] = sport["teamMap"][participants[0]]
            if participants[1] in sport["teamMap"]:
              matches[-1]["teams"][1]["name"] = sport["teamMap"][participants[1]]

      if matches:
        sport["stages"].append({"name": stage_label, "matches": matches})

      row_index = section_end

  sports = list(sports_map.values())
  sports.sort(key=lambda sport: sport.get("name", ""))

  max_day = 1
  max_g = 0
  for sport in sports:
    for stage in sport.get("stages", []):
      for match in stage.get("matches", []):
        day_value = match.get("day")
        if isinstance(day_value, int) and day_value > max_day:
          max_day = day_value
        slot_text = str(match.get("slot") or "")
        for part in slot_text.split("+"):
          cleaned = part.strip().upper().replace("G", "")
          if cleaned.isdigit():
            max_g = max(max_g, int(cleaned))

  return {
    "meta": {
      "title": f"{grade_tag.upper()} 赛程",
      "startDate": "",
      "dayMap": build_day_map(max_day),
      "timeSlots": {"default": build_default_time_slots(max(7, max_g))},
      "updatedAt": ""
    },
    "sports": sports
  }


def build_scores(schedule_data):
  matches = {}
  for sport in schedule_data.get("sports", []):
    for stage in sport.get("stages", []):
      for match in stage.get("matches", []):
        matches[match["id"]] = {
          "scoreA": None,
          "scoreB": None,
          "status": "scheduled",
          "note": ""
        }
  return {"meta": {"updatedAt": ""}, "matches": matches}


def main():
  schedule_high = json.loads((DATA / "schedule.json").read_text(encoding="utf-8"))
  scores_high = json.loads((DATA / "scores.json").read_text(encoding="utf-8"))

  (DATA / "schedule-high.json").write_text(
    json.dumps(schedule_high, ensure_ascii=False, indent=2), encoding="utf-8"
  )
  (DATA / "scores-high.json").write_text(
    json.dumps(scores_high, ensure_ascii=False, indent=2), encoding="utf-8"
  )

  schedule_j1 = parse_workbook(ROOT / "初一赛程表.xlsx", "j1")
  schedule_j2 = parse_workbook(ROOT / "初二赛程表.xlsx", "j2")

  (DATA / "schedule-j1.json").write_text(
    json.dumps(schedule_j1, ensure_ascii=False, indent=2), encoding="utf-8"
  )
  (DATA / "schedule-j2.json").write_text(
    json.dumps(schedule_j2, ensure_ascii=False, indent=2), encoding="utf-8"
  )
  (DATA / "scores-j1.json").write_text(
    json.dumps(build_scores(schedule_j1), ensure_ascii=False, indent=2), encoding="utf-8"
  )
  (DATA / "scores-j2.json").write_text(
    json.dumps(build_scores(schedule_j2), ensure_ascii=False, indent=2), encoding="utf-8"
  )

  schedule_prep = {
    "meta": {
      "title": "PREP 赛程",
      "startDate": "",
      "dayMap": {},
      "timeSlots": {"default": {}},
      "updatedAt": "待体育部发布"
    },
    "sports": []
  }
  scores_prep = {"meta": {"updatedAt": "待体育部发布"}, "matches": {}}

  (DATA / "schedule-prep.json").write_text(
    json.dumps(schedule_prep, ensure_ascii=False, indent=2), encoding="utf-8"
  )
  (DATA / "scores-prep.json").write_text(
    json.dumps(scores_prep, ensure_ascii=False, indent=2), encoding="utf-8"
  )

  classes = []
  for index in range(1, 11):
    classes.append({"key": f"中预{index}", "label": f"中预 {index} 班", "group": "prep"})
  for index in range(1, 9):
    classes.append({"key": f"初一{index}", "label": f"初一 {index} 班", "group": "j1"})
  for index in range(1, 9):
    classes.append({"key": f"初二{index}", "label": f"初二 {index} 班", "group": "j2"})
  for index in range(1, 7):
    classes.append({"key": f"高一{index}", "label": f"高一 {index} 班", "group": "high"})
  classes.append({"key": "高一VCE", "label": "高一 VCE", "group": "high"})
  for index in range(1, 7):
    classes.append({"key": f"高二{index}", "label": f"高二 {index} 班", "group": "high"})
  classes.append({"key": "高二VCE", "label": "高二 VCE", "group": "high"})

  class_config = {
    "defaultClass": "高一1",
    "groups": {
      "prep": {
        "name": "中预",
        "schedule": ["data/schedule-prep.json"],
        "scores": ["data/scores-prep.json"]
      },
      "j1": {
        "name": "初一",
        "schedule": ["data/schedule-j1.json"],
        "scores": ["data/scores-j1.json"]
      },
      "j2": {
        "name": "初二",
        "schedule": ["data/schedule-j2.json"],
        "scores": ["data/scores-j2.json"]
      },
      "high": {
        "name": "高中",
        "schedule": ["data/schedule-high.json"],
        "scores": ["data/scores-high.json"]
      }
    },
    "classes": classes
  }

  (DATA / "class-config.json").write_text(
    json.dumps(class_config, ensure_ascii=False, indent=2), encoding="utf-8"
  )

  print("Generated split schedule/scores/class-config files.")


if __name__ == "__main__":
  main()
